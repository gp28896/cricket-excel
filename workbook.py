"""
workbook.py
===========

Central workbook orchestration module for the Cricket Tournament Excel system.

Overview
--------
This module acts as the highest-level coordinator responsible for constructing,
configuring, and exporting complete Excel workbooks that represent an entire
cricket tournament.

Unlike lower-level modules that focus on a single concern (such as formulas,
validation, constants, styling, or worksheet generation), this module provides
the orchestration layer that combines those components into a fully functional
spreadsheet application.

The responsibilities of this module include (implemented incrementally across
future parts of this file):

* Workbook creation
* Workbook configuration
* Metadata initialization
* Worksheet orchestration
* Named range registration
* Formula integration
* Data validation registration
* Conditional formatting integration
* Freeze panes
* Print settings
* Workbook properties
* Protection configuration
* Export helpers
* Version compatibility checks
* Logging
* High-level error handling

Design Philosophy
-----------------
The architecture intentionally separates orchestration from implementation.

Each worksheet should be created by a dedicated builder or helper rather than
embedding worksheet logic directly inside this module. Likewise, formatting,
validation, formulas, and constants are delegated to their respective modules.

This design provides:

* excellent maintainability
* easier testing
* modular development
* predictable extension points
* reduced coupling
* simpler future refactoring

Thread Safety
-------------
Workbook objects provided by openpyxl are mutable and should not be considered
thread-safe. A workbook should be constructed and modified from a single thread.

Python Version
--------------
Python 3.12+

Third-party Dependencies
------------------------
* openpyxl

Internal Dependencies
---------------------
* constants
* formulas
* validation

Nothing in this initial module foundation performs workbook creation. This file
currently establishes the shared imports, typing infrastructure, logging, and
module-level constants used throughout later sections.
"""

from __future__ import annotations

###############################################################################
# Standard library imports
###############################################################################

import logging
from pathlib import Path
from typing import Final
from typing import TypeAlias
from datetime import UTC, datetime
from collections.abc import Iterable, Sequence

from pathlib import Path
from typing import BinaryIO, Final, Self

###############################################################################
# Third-party imports
###############################################################################

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Protection
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from contextlib import contextmanager
from time import perf_counter

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.cell import rows_from_range


###############################################################################
# Internal project imports
###############################################################################

import constants
import formulas
import validation
from dataclasses import dataclass, field

###############################################################################
# Type aliases
###############################################################################

WorkbookType: TypeAlias = Workbook
"""Alias representing an openpyxl workbook."""

WorksheetType: TypeAlias = Worksheet
"""Alias representing an openpyxl worksheet."""

CellType: TypeAlias = Cell
"""Alias representing an openpyxl cell."""

PathLike: TypeAlias = str | Path
"""
Represents an output path accepted by workbook export helpers.

The alias allows callers to supply either a string path or a pathlib.Path
instance.
"""

###############################################################################
# Logging configuration
###############################################################################

LOGGER_NAME: Final[str] = "cricket_excel.workbook"
"""
Canonical logger name used throughout this module.

Applications embedding this project may configure this logger independently
from the root logger to obtain detailed workbook construction diagnostics.
"""

logger = logging.getLogger(LOGGER_NAME)

###############################################################################
# Module constants
###############################################################################

MODULE_NAME: Final[str] = "workbook"
"""Human-readable module identifier."""

MODULE_VERSION: Final[str] = "1.0.0"
"""
Internal module implementation version.

This value identifies the implementation of this orchestration module rather
than the version of the overall application.
"""

DEFAULT_SHEET_INDEX: Final[int] = 0
"""
Default insertion index for worksheets.

Future worksheet builders may override this value when constructing workbooks
with custom ordering.
"""

EXCEL_MAX_ROWS: Final[int] = 1_048_576
"""Maximum number of rows supported by modern Excel workbooks."""

EXCEL_MAX_COLUMNS: Final[int] = 16_384
"""Maximum number of columns supported by modern Excel workbooks."""

EXCEL_FIRST_ROW: Final[int] = 1
"""Index of the first worksheet row."""

EXCEL_FIRST_COLUMN: Final[int] = 1
"""Index of the first worksheet column."""

###############################################################################
# Internal helper constants
###############################################################################

_INTERNAL_PREFIX: Final[str] = "_"
"""
Naming prefix reserved for workbook-internal identifiers.

Future named ranges, hidden worksheets, and implementation-specific resources
may use this prefix to distinguish them from user-visible objects.
"""

_LOG_MESSAGE_PREFIX: Final[str] = "[Workbook]"
"""
Standard prefix applied to diagnostic log messages generated by this module.
"""

_COLUMN_CACHE_LIMIT: Final[int] = EXCEL_MAX_COLUMNS
"""
Upper bound used by future column-letter helper caches.

The value matches Excel's maximum supported column count.
"""

###############################################################################
# Public exports
###############################################################################

__all__: list[str] = []

###############################################################################
# Workbook configuration
###############################################################################




@dataclass(slots=True, kw_only=True)
class WorkbookConfig:
    """Configuration used by :class:`WorkbookBuilder`.

    This dataclass centralizes every configurable option required to construct
    a workbook. Future parts of ``workbook.py`` will consume this object during
    workbook creation, worksheet initialization, styling, printing, protection,
    and export.

    Keeping all configuration in a single immutable-style object makes the
    workbook builder easier to test, extend, and validate.

    Attributes:
        workbook_title:
            Human-readable workbook title.

        author:
            Workbook author written to document properties.

        company:
            Company or organization associated with the workbook.

        tournament_name:
            Name of the cricket tournament represented by the workbook.

        default_font:
            Default font family applied throughout the workbook.

        default_font_size:
            Default font size in points.

        default_row_height:
            Default worksheet row height.

        default_column_width:
            Default worksheet column width.

        theme_name:
            Logical theme identifier. Future style modules may map this value
            to fonts, colours, borders, and fills.

        date_format:
            Excel number format for dates.

        time_format:
            Excel number format for times.

        datetime_format:
            Excel number format for combined date/time values.

        workbook_protection_enabled:
            Enables workbook-level protection by default.

        structure_protection:
            Protect workbook structure.

        windows_protection:
            Protect workbook window layout.

        default_sheet_protection:
            Enable worksheet protection by default.

        allow_select_locked_cells:
            Allow selecting locked cells.

        allow_select_unlocked_cells:
            Allow selecting unlocked cells.

        fit_to_page:
            Enable worksheet fit-to-page printing.

        fit_to_width:
            Printed page width.

        fit_to_height:
            Printed page height.

        landscape:
            Print using landscape orientation.

        center_horizontally:
            Horizontally center printed pages.

        center_vertically:
            Vertically center printed pages.
    """

    workbook_title: str = "Cricket Tournament Workbook"
    author: str = "Cricket Tournament Manager"
    company: str = ""
    tournament_name: str = ""

    default_font: str = "Calibri"
    default_font_size: float = 11.0

    default_row_height: float = 15.0
    default_column_width: float = 12.0

    theme_name: str = "default"

    date_format: str = "yyyy-mm-dd"
    time_format: str = "hh:mm:ss"
    datetime_format: str = "yyyy-mm-dd hh:mm:ss"

    workbook_protection_enabled: bool = False
    structure_protection: bool = False
    windows_protection: bool = False

    default_sheet_protection: bool = False
    allow_select_locked_cells: bool = True
    allow_select_unlocked_cells: bool = True

    fit_to_page: bool = True
    fit_to_width: int = 1
    fit_to_height: int = 0
    landscape: bool = True
    center_horizontally: bool = True
    center_vertically: bool = False

    def __post_init__(self) -> None:
        """Validate configuration values.

        Raises:
            TypeError:
                If a value has an unexpected type.

            ValueError:
                If a value falls outside the supported range.
        """
        text_fields = (
            ("workbook_title", self.workbook_title),
            ("author", self.author),
            ("company", self.company),
            ("tournament_name", self.tournament_name),
            ("default_font", self.default_font),
            ("theme_name", self.theme_name),
            ("date_format", self.date_format),
            ("time_format", self.time_format),
            ("datetime_format", self.datetime_format),
        )

        for name, value in text_fields:
            if not isinstance(value, str):
                raise TypeError(f"{name} must be a string.")

            if not value.strip() and name not in {"company", "tournament_name"}:
                raise ValueError(f"{name} cannot be empty.")

        if self.default_font_size <= 0:
            raise ValueError("default_font_size must be greater than zero.")

        if self.default_row_height <= 0:
            raise ValueError("default_row_height must be greater than zero.")

        if self.default_column_width <= 0:
            raise ValueError("default_column_width must be greater than zero.")

        if self.fit_to_width < 0:
            raise ValueError("fit_to_width cannot be negative.")

        if self.fit_to_height < 0:
            raise ValueError("fit_to_height cannot be negative.")


###############################################################################
# Workbook builder skeleton
###############################################################################


class WorkbookBuilder:
    """High-level workbook orchestration class.

    The builder coordinates construction of the complete Excel workbook while
    delegating specialized responsibilities to lower-level modules.

    Only the foundational object model is established here. Workbook creation,
    worksheet creation, styling, validation, formulas, named ranges, printing,
    protection, and export are implemented in later parts of this module.
    """

    def __init__(
        self,
        config: WorkbookConfig | None = None,
    ) -> None:
        """Initialize the workbook builder.

        Args:
            config:
                Optional workbook configuration. When omitted, a default
                configuration is created.
        """
        self._config: WorkbookConfig = config or WorkbookConfig()

        self._workbook: WorkbookType | None = None
        self._worksheet_registry: dict[str, WorksheetType] = {}

        self._logger: logging.Logger = logger

        self._initialize_helpers()

    @property
    def config(self) -> WorkbookConfig:
        """Return the active workbook configuration."""
        return self._config

    @property
    def workbook(self) -> WorkbookType | None:
        """Return the managed workbook instance.

        Returns:
            The workbook once created, otherwise ``None``.
        """
        return self._workbook

    @property
    def worksheet_registry(self) -> dict[str, WorksheetType]:
        """Return the internal worksheet registry.

        Future workbook construction stages will populate this mapping with
        logical worksheet names and their corresponding worksheet objects.
        """
        return self._worksheet_registry

    @property
    def logger(self) -> logging.Logger:
        """Return the module logger."""
        return self._logger

    def _initialize_helpers(self) -> None:
        """Initialize internal helper infrastructure.

        The initial implementation prepares extension points required by later
        workbook construction stages. Future parts may populate caches,
        registries, style managers, validation managers, formula helpers, and
        workbook services.
        """
        self._helper_registry: dict[str, object] = {}



###############################################################################
# WorkbookBuilder - workbook lifecycle methods
###############################################################################

    def create_workbook(self) -> WorkbookType:
        """Create and initialize a new workbook.

        This method is the primary entry point for workbook construction.
        Individual initialization steps are intentionally delegated to smaller
        methods so subclasses may override specific phases without replacing the
        complete workflow.

        The current implementation performs only foundational initialization.
        Worksheet population, styling, formulas, validations, and other
        application-specific features are implemented in later parts of this
        module.

        Returns:
            The newly created workbook.

        Raises:
            RuntimeError:
                If workbook creation fails.
        """
        self._logger.info("%s Creating workbook.", _LOG_MESSAGE_PREFIX)

        try:
            self._workbook = Workbook()

            self.initialize_internal_state()
            self.initialize_workbook()
            self.remove_default_sheet()
            self.set_workbook_properties()
            self.create_core_sheets()
            self.create_hidden_sheets()

            self._logger.info(
                "%s Workbook initialization completed successfully.",
                _LOG_MESSAGE_PREFIX,
            )

            return self._workbook

        except Exception as exc:
            self._logger.exception(
                "%s Workbook creation failed.",
                _LOG_MESSAGE_PREFIX,
            )
            self._workbook = None
            raise RuntimeError("Unable to create workbook.") from exc

    def initialize_workbook(self) -> None:
        """Initialize workbook-wide settings.

        This method performs only high-level initialization that does not depend
        upon worksheet creation.

        Future implementations may configure workbook calculation mode,
        workbook security, themes, custom properties, named styles, and other
        workbook-wide resources.

        Raises:
            RuntimeError:
                If no workbook exists.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Initializing workbook.",
            _LOG_MESSAGE_PREFIX,
        )

    def remove_default_sheet(self) -> None:
        """Remove the automatically created worksheet.

        ``openpyxl`` creates a worksheet named ``Sheet`` whenever a workbook is
        instantiated. This application manages worksheet creation explicitly, so
        the automatically generated sheet is removed before custom worksheets
        are added.

        Raises:
            RuntimeError:
                If the workbook has not yet been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Removing default worksheet if present.",
            _LOG_MESSAGE_PREFIX,
        )

        try:
            if len(self._workbook.worksheets) != 1:
                return

            default_sheet = self._workbook.active

            if default_sheet is not None:
                self._workbook.remove(default_sheet)

        except Exception:
            self._logger.exception(
                "%s Failed while removing default worksheet.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

    def set_workbook_properties(self) -> None:
        """Populate workbook document properties.

        Metadata originates from :class:`WorkbookConfig`.

        Future implementations may also configure custom document properties,
        categories, keywords, revision information, and application metadata.

        Raises:
            RuntimeError:
                If no workbook exists.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Applying workbook properties.",
            _LOG_MESSAGE_PREFIX,
        )

        try:
            properties = self._workbook.properties

            properties.title = self._config.workbook_title
            properties.creator = self._config.author
            properties.lastModifiedBy = self._config.author
            properties.company = self._config.company
            properties.subject = self._config.tournament_name
            properties.category = "Cricket Tournament"

        except Exception:
            self._logger.exception(
                "%s Failed to apply workbook properties.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

    def create_core_sheets(self) -> None:
        """Create all visible worksheets.

        The foundational implementation intentionally performs no worksheet
        creation. Future parts of this module will construct every visible
        worksheet in the required order and register each sheet inside the
        worksheet registry.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Core worksheet creation phase started.",
            _LOG_MESSAGE_PREFIX,
        )

    def create_hidden_sheets(self) -> None:
        """Create internal worksheets.

        Hidden worksheets will later store lookup values, validation lists,
        configuration data, and other internal resources required by the
        workbook.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Hidden worksheet creation phase started.",
            _LOG_MESSAGE_PREFIX,
        )

    def initialize_internal_state(self) -> None:
        """Reset internal builder state.

        This method prepares internal registries and caches before workbook
        construction begins. Calling it allows a single builder instance to be
        safely reused for constructing multiple workbooks sequentially.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Initializing internal state.",
            _LOG_MESSAGE_PREFIX,
        )

        self._worksheet_registry.clear()

        helper_registry = getattr(self, "_helper_registry", None)
        if isinstance(helper_registry, dict):
            helper_registry.clear()


###############################################################################
# WorkbookBuilder - style initialization
###############################################################################

    def initialize_styles(self) -> None:
        """Initialize workbook styling infrastructure.

        This method prepares all styling resources required by the workbook.

        The implementation intentionally delegates styling responsibilities to
        the project's :mod:`styles` module rather than embedding formatting
        logic directly inside :class:`WorkbookBuilder`. This keeps workbook
        orchestration independent from visual presentation.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            ImportError:
                If the project's ``styles`` module cannot be imported.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Initializing workbook styles.",
            _LOG_MESSAGE_PREFIX,
        )

        try:
            import styles
        except ImportError:
            self._logger.exception(
                "%s Unable to import styles module.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

        self._helper_registry["styles"] = styles

        self.register_named_styles()

    def register_named_styles(self) -> None:
        """Register workbook named styles.

        Registration is delegated to ``styles.py`` whenever a compatible
        registration function is available.

        Supported entry points (checked in order):

        * ``register_named_styles(workbook)``
        * ``initialize_styles(workbook)``

        If neither entry point exists, this method simply records the condition
        in the log without failing. This allows the styling module to evolve
        independently while maintaining backward compatibility.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        self._logger.debug(
            "%s Registering named styles.",
            _LOG_MESSAGE_PREFIX,
        )

        register = getattr(styles_module, "register_named_styles", None)

        if callable(register):
            register(self._workbook)
            return

        initialize = getattr(styles_module, "initialize_styles", None)

        if callable(initialize):
            initialize(self._workbook)
            return

        self._logger.info(
            "%s No named-style registration function found.",
            _LOG_MESSAGE_PREFIX,
        )

    def apply_default_style(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Apply the project's default worksheet style.

        Args:
            worksheet:
                Worksheet receiving the default appearance.

        Styling is delegated to ``styles.py`` using one of the following
        supported entry points:

        * ``apply_default_style(worksheet, config)``
        * ``apply_default_style(worksheet)``
        * ``style_worksheet(worksheet)``
        """
        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        self._logger.debug(
            "%s Applying default style to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

        apply = getattr(styles_module, "apply_default_style", None)

        if callable(apply):
            try:
                apply(worksheet, self._config)
            except TypeError:
                apply(worksheet)
            return

        worksheet_style = getattr(styles_module, "style_worksheet", None)

        if callable(worksheet_style):
            worksheet_style(worksheet)
            return

        self._logger.info(
            "%s No worksheet default style function available.",
            _LOG_MESSAGE_PREFIX,
        )

    def style_cell(
        self,
        cell: CellType,
        style_name: str,
    ) -> None:
        """Apply a named style to a single cell.

        Args:
            cell:
                Target cell.

            style_name:
                Logical style identifier defined by ``styles.py``.
        """
        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        styler = getattr(styles_module, "style_cell", None)

        if not callable(styler):
            raise AttributeError(
                "styles.style_cell() is not available."
            )

        styler(cell, style_name)

    def style_range(
        self,
        worksheet: WorksheetType,
        cell_range: str,
        style_name: str,
    ) -> None:
        """Apply a named style to a rectangular cell range.

        Args:
            worksheet:
                Target worksheet.

            cell_range:
                Excel range notation.

            style_name:
                Logical style identifier.
        """
        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        styler = getattr(styles_module, "style_range", None)

        if not callable(styler):
            raise AttributeError(
                "styles.style_range() is not available."
            )

        styler(worksheet, cell_range, style_name)

    def style_row(
        self,
        worksheet: WorksheetType,
        row_index: int,
        style_name: str,
    ) -> None:
        """Apply a named style to an entire worksheet row.

        Args:
            worksheet:
                Target worksheet.

            row_index:
                One-based Excel row number.

            style_name:
                Logical style identifier.
        """
        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        styler = getattr(styles_module, "style_row", None)

        if not callable(styler):
            raise AttributeError(
                "styles.style_row() is not available."
            )

        styler(worksheet, row_index, style_name)

    def style_column(
        self,
        worksheet: WorksheetType,
        column: int | str,
        style_name: str,
    ) -> None:
        """Apply a named style to an entire worksheet column.

        Args:
            worksheet:
                Target worksheet.

            column:
                Excel column letter or one-based column index.

            style_name:
                Logical style identifier.
        """
        styles_module = self._helper_registry.get("styles")

        if styles_module is None:
            raise RuntimeError(
                "Style subsystem has not been initialized."
            )

        styler = getattr(styles_module, "style_column", None)

        if not callable(styler):
            raise AttributeError(
                "styles.style_column() is not available."
            )

        styler(worksheet, column, style_name)


###############################################################################
# WorkbookBuilder - Theme support
###############################################################################

    def apply_theme(self) -> None:
        """Apply the configured workbook theme.

        This method prepares workbook-wide visual defaults used by later
        worksheet builders. It does not directly format worksheets or cells.
        Instead, it creates a centralized theme definition that other builder
        methods and the project's ``styles`` module can consume.

        Theme application is intentionally lightweight:

        * validates the configured theme
        * integrates with ``constants.py`` where available
        * reuses ``styles.py`` for theme registration if supported
        * prepares cached default style definitions

        Raises:
            RuntimeError:
                If the workbook has not yet been created.

            ValueError:
                If the configured theme is invalid.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Applying workbook theme '%s'.",
            _LOG_MESSAGE_PREFIX,
            self._config.theme_name,
        )

        theme_name = self._config.theme_name.strip()

        if not theme_name:
            raise ValueError("Theme name cannot be empty.")

        theme: dict[str, object] = {
            "name": theme_name,
            "colors": self._theme_color_mapping(),
            "fonts": self._theme_font_defaults(),
            "borders": self._theme_border_defaults(),
            "fills": self._theme_fill_defaults(),
            "alignment": self._theme_alignment_defaults(),
            "number_formats": self._theme_number_format_defaults(),
        }

        self._helper_registry["theme"] = theme

        styles_module = self._helper_registry.get("styles")

        if styles_module is not None:
            register = getattr(styles_module, "apply_theme", None)

            if callable(register):
                try:
                    register(self._workbook, theme)
                except TypeError:
                    register(theme)

        self._logger.info(
            "%s Theme '%s' initialized.",
            _LOG_MESSAGE_PREFIX,
            theme_name,
        )

    def _theme_color_mapping(self) -> dict[str, str]:
        """Build the workbook theme colour mapping.

        The method first attempts to reuse colour definitions exposed by
        ``constants.py``. If no compatible mapping exists, conservative Excel
        defaults are used.

        Returns:
            Mapping of logical colour names to RGB values.
        """
        mapping = getattr(constants, "THEME_COLORS", None)

        if isinstance(mapping, dict):
            return dict(mapping)

        return {
            "primary": "1F4E78",
            "secondary": "D9EAD3",
            "accent": "4F81BD",
            "success": "70AD47",
            "warning": "FFC000",
            "danger": "C00000",
            "header_background": "D9E2F3",
            "header_text": "000000",
            "grid": "D9D9D9",
            "text": "000000",
            "background": "FFFFFF",
        }

    def _theme_font_defaults(self) -> dict[str, object]:
        """Create workbook font defaults.

        Returns:
            Dictionary describing default workbook font settings.
        """
        return {
            "family": self._config.default_font,
            "size": self._config.default_font_size,
            "bold": False,
            "italic": False,
            "underline": None,
            "strike": False,
            "color": "000000",
        }

    def _theme_border_defaults(self) -> dict[str, object]:
        """Create workbook border defaults.

        Returns:
            Border configuration consumed by later styling phases.
        """
        return {
            "style": "thin",
            "color": "D9D9D9",
            "outline": True,
        }

    def _theme_fill_defaults(self) -> dict[str, object]:
        """Create workbook fill defaults.

        Returns:
            Fill configuration for reusable workbook styles.
        """
        return {
            "pattern": "solid",
            "foreground": "FFFFFF",
            "background": "FFFFFF",
        }

    def _theme_alignment_defaults(self) -> dict[str, object]:
        """Create workbook alignment defaults.

        Returns:
            Alignment configuration shared throughout the workbook.
        """
        return {
            "horizontal": "center",
            "vertical": "center",
            "wrap_text": True,
            "shrink_to_fit": False,
        }

    def _theme_number_format_defaults(self) -> dict[str, str]:
        """Create workbook number-format defaults.

        Returns:
            Mapping of logical format names to Excel number formats.
        """
        return {
            "date": self._config.date_format,
            "time": self._config.time_format,
            "datetime": self._config.datetime_format,
            "integer": "0",
            "decimal": "0.00",
            "percentage": "0.00%",
            "currency": "#,##0.00",
            "text": "@",
        }


###############################################################################
# WorkbookBuilder - Named style registration
###############################################################################




    def register_header_style(self) -> NamedStyle:
        """Register the standard header style.

        The style is registered only once per workbook. Subsequent calls return
        the existing style instance.

        Returns:
            The registered :class:`~openpyxl.styles.NamedStyle`.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        return self._register_named_style(
            style_name="Header",
            number_format="General",
            style_callback="create_header_named_style",
        )

    def register_title_style(self) -> NamedStyle:
        """Register the workbook title style.

        Returns:
            The registered title style.
        """
        return self._register_named_style(
            style_name="Title",
            number_format="General",
            style_callback="create_title_named_style",
        )

    def register_normal_style(self) -> NamedStyle:
        """Register the default workbook style.

        Returns:
            The registered normal style.
        """
        return self._register_named_style(
            style_name="Normal",
            number_format="General",
            style_callback="create_normal_named_style",
        )

    def register_number_style(self) -> NamedStyle:
        """Register the standard numeric style.

        Returns:
            The registered numeric style.
        """
        return self._register_named_style(
            style_name="Number",
            number_format=self._theme_number_format_defaults()["integer"],
            style_callback="create_number_named_style",
        )

    def register_percentage_style(self) -> NamedStyle:
        """Register the percentage style.

        Returns:
            The registered percentage style.
        """
        return self._register_named_style(
            style_name="Percentage",
            number_format=self._theme_number_format_defaults()["percentage"],
            style_callback="create_percentage_named_style",
        )

    def register_date_style(self) -> NamedStyle:
        """Register the date style.

        Returns:
            The registered date style.
        """
        return self._register_named_style(
            style_name="Date",
            number_format=self._config.date_format,
            style_callback="create_date_named_style",
        )

    def register_time_style(self) -> NamedStyle:
        """Register the time style.

        Returns:
            The registered time style.
        """
        return self._register_named_style(
            style_name="Time",
            number_format=self._config.time_format,
            style_callback="create_time_named_style",
        )

    def register_currency_style(self) -> NamedStyle:
        """Register the currency style.

        Returns:
            The registered currency style.
        """
        return self._register_named_style(
            style_name="Currency",
            number_format=self._theme_number_format_defaults()["currency"],
            style_callback="create_currency_named_style",
        )

    def register_hidden_style(self) -> NamedStyle:
        """Register the hidden/internal worksheet style.

        Returns:
            The registered hidden style.
        """
        return self._register_named_style(
            style_name="Hidden",
            number_format="General",
            style_callback="create_hidden_named_style",
        )

    def _register_named_style(
        self,
        *,
        style_name: str,
        number_format: str,
        style_callback: str,
    ) -> NamedStyle:
        """Register a named style if it does not already exist.

        The method first attempts to reuse style factory functions from
        ``styles.py``. If no compatible factory exists, a minimal
        :class:`~openpyxl.styles.NamedStyle` is created using the active workbook
        theme defaults.

        Args:
            style_name:
                Excel named style name.

            number_format:
                Default Excel number format.

            style_callback:
                Factory function expected inside ``styles.py``.

        Returns:
            The registered named style.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        for existing in self._workbook.named_styles:
            if isinstance(existing, NamedStyle):
                if existing.name == style_name:
                    return existing
            elif existing == style_name:
                for candidate in self._workbook._named_styles:
                    if candidate.name == style_name:
                        return candidate

        styles_module = self._helper_registry.get("styles")

        style: NamedStyle | None = None

        if styles_module is not None:
            factory = getattr(styles_module, style_callback, None)

            if callable(factory):
                try:
                    style = factory(self._config)
                except TypeError:
                    style = factory()

        if style is None:
            theme = self._helper_registry.get("theme", {})

            fonts = theme.get("fonts", {})
            borders = theme.get("borders", {})
            fills = theme.get("fills", {})
            alignment = theme.get("alignment", {})

            style = NamedStyle(name=style_name)

            style.font = Font(
                name=fonts.get("family", self._config.default_font),
                size=fonts.get("size", self._config.default_font_size),
                bold=fonts.get("bold", False),
                italic=fonts.get("italic", False),
                color=fonts.get("color", "000000"),
            )

            style.fill = PatternFill(
                fill_type=fills.get("pattern", "solid"),
                fgColor=fills.get("foreground", "FFFFFF"),
                bgColor=fills.get("background", "FFFFFF"),
            )

            side = Side(
                style=borders.get("style", "thin"),
                color=borders.get("color", "D9D9D9"),
            )

            style.border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side,
            )

            style.alignment = Alignment(
                horizontal=alignment.get("horizontal", "center"),
                vertical=alignment.get("vertical", "center"),
                wrap_text=alignment.get("wrap_text", True),
                shrink_to_fit=alignment.get("shrink_to_fit", False),
            )

            style.number_format = number_format

        if style.name != style_name:
            style.name = style_name

        self._workbook.add_named_style(style)

        self._logger.debug(
            "%s Registered named style '%s'.",
            _LOG_MESSAGE_PREFIX,
            style_name,
        )

        return style


###############################################################################
# WorkbookBuilder - Reusable column utilities
###############################################################################

    def set_column_width(
        self,
        worksheet: WorksheetType,
        column: int | str,
        width: float,
    ) -> None:
        """Set the width of a single worksheet column.

        Args:
            worksheet:
                Target worksheet.

            column:
                Excel column reference. May be either a one-based integer column
                index or an Excel column letter.

            width:
                Desired column width.

        Raises:
            ValueError:
                If the column reference or width is invalid.
        """
        column_letter = self.validate_column_reference(column)

        if width <= 0:
            raise ValueError("Column width must be greater than zero.")

        worksheet.column_dimensions[column_letter].width = width

        self._logger.debug(
            "%s Set width of column '%s' to %.2f.",
            _LOG_MESSAGE_PREFIX,
            column_letter,
            width,
        )

    def set_multiple_column_widths(
        self,
        worksheet: WorksheetType,
        widths: dict[int | str, float],
    ) -> None:
        """Apply widths to multiple worksheet columns.

        Args:
            worksheet:
                Target worksheet.

            widths:
                Mapping of column reference to width.
        """
        for column, width in widths.items():
            self.set_column_width(
                worksheet=worksheet,
                column=column,
                width=width,
            )

    def autosize_columns(
        self,
        worksheet: WorksheetType,
        *,
        minimum_width: float = 8.43,
        maximum_width: float = 60.0,
        padding: float = 2.0,
    ) -> None:
        """Automatically size populated worksheet columns.

        The calculation is based on the string representation of each cell
        value. This helper intentionally provides a generic implementation and
        does not contain worksheet-specific assumptions.

        Args:
            worksheet:
                Target worksheet.

            minimum_width:
                Minimum permitted column width.

            maximum_width:
                Maximum permitted column width.

            padding:
                Extra width added after measuring the largest cell value.
        """
        self._logger.debug(
            "%s Autosizing worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

        for column_cells in worksheet.columns:
            try:
                first_cell = next(iter(column_cells))
            except StopIteration:
                continue

            column_letter = first_cell.column_letter

            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

            calculated_width = max(
                minimum_width,
                min(maximum_width, max_length + padding),
            )

            worksheet.column_dimensions[column_letter].width = calculated_width

    def hide_columns(
        self,
        worksheet: WorksheetType,
        *columns: int | str,
    ) -> None:
        """Hide one or more worksheet columns.

        Args:
            worksheet:
                Target worksheet.

            *columns:
                Column references.
        """
        for column in columns:
            column_letter = self.validate_column_reference(column)
            worksheet.column_dimensions[column_letter].hidden = True

    def unhide_columns(
        self,
        worksheet: WorksheetType,
        *columns: int | str,
    ) -> None:
        """Unhide one or more worksheet columns.

        Args:
            worksheet:
                Target worksheet.

            *columns:
                Column references.
        """
        for column in columns:
            column_letter = self.validate_column_reference(column)
            worksheet.column_dimensions[column_letter].hidden = False

    def group_columns(
        self,
        worksheet: WorksheetType,
        start_column: int | str,
        end_column: int | str,
        *,
        hidden: bool = False,
        outline_level: int = 1,
    ) -> None:
        """Create an Excel outline group for a range of columns.

        Args:
            worksheet:
                Target worksheet.

            start_column:
                First column in the group.

            end_column:
                Last column in the group.

            hidden:
                Whether the grouped columns should initially be hidden.

            outline_level:
                Excel outline level.

        Raises:
            ValueError:
                If the outline level is invalid.
        """
        if outline_level < 1:
            raise ValueError("outline_level must be at least 1.")

        start = self.validate_column_reference(start_column)
        end = self.validate_column_reference(end_column)

        worksheet.column_dimensions.group(
            start=start,
            end=end,
            hidden=hidden,
            outline_level=outline_level,
        )

    def copy_column_widths(
        self,
        source: WorksheetType,
        destination: WorksheetType,
    ) -> None:
        """Copy all explicit column widths between worksheets.

        Args:
            source:
                Worksheet providing the widths.

            destination:
                Worksheet receiving the widths.
        """
        for column_letter, dimension in source.column_dimensions.items():
            destination.column_dimensions[column_letter].width = dimension.width
            destination.column_dimensions[column_letter].hidden = (
                dimension.hidden
            )
            destination.column_dimensions[column_letter].bestFit = (
                dimension.bestFit
            )
            destination.column_dimensions[column_letter].outlineLevel = (
                dimension.outlineLevel
            )

        self._logger.debug(
            "%s Copied column widths from '%s' to '%s'.",
            _LOG_MESSAGE_PREFIX,
            source.title,
            destination.title,
        )

    def validate_column_reference(
        self,
        column: int | str,
    ) -> str:
        """Validate and normalize a column reference.

        Accepted formats are:

        * one-based integer column indices
        * Excel column letters (for example ``A`` or ``AA``)

        Args:
            column:
                Column reference.

        Returns:
            Normalized uppercase Excel column letter.

        Raises:
            TypeError:
                If the supplied type is unsupported.

            ValueError:
                If the reference is outside Excel's supported limits.
        """
        if isinstance(column, int):
            if not (1 <= column <= EXCEL_MAX_COLUMNS):
                raise ValueError(
                    f"Column index must be between 1 and {EXCEL_MAX_COLUMNS}."
                )

            return get_column_letter(column)

        if isinstance(column, str):
            normalized = column.strip().upper()

            if not normalized.isalpha():
                raise ValueError(
                    "Column reference must contain only letters."
                )

            return normalized

        raise TypeError(
            "Column reference must be an integer or Excel column letter."
        )


###############################################################################
# WorkbookBuilder - Worksheet management
###############################################################################

    def create_sheet(
        self,
        title: str,
        *,
        index: int | None = None,
    ) -> WorksheetType:
        """Create and register a worksheet.

        The worksheet registry maintained by :class:`WorkbookBuilder` is treated
        as the authoritative source for workbook worksheets. Every worksheet
        created through this method is automatically registered.

        Args:
            title:
                Worksheet title.

            index:
                Optional insertion index.

        Returns:
            The newly created worksheet.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            ValueError:
                If the worksheet title is invalid or already exists.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        title = title.strip()

        if not title:
            raise ValueError("Worksheet title cannot be empty.")

        if self.sheet_exists(title):
            raise ValueError(f"Worksheet '{title}' already exists.")

        worksheet = self._workbook.create_sheet(
            title=title,
            index=index,
        )

        self._worksheet_registry[title] = worksheet

        self._logger.info(
            "%s Created worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

        return worksheet

    def get_sheet(
        self,
        title: str,
    ) -> WorksheetType:
        """Return a registered worksheet.

        Args:
            title:
                Worksheet title.

        Returns:
            Registered worksheet.

        Raises:
            KeyError:
                If the worksheet is not registered.
        """
        try:
            return self._worksheet_registry[title]
        except KeyError as exc:
            raise KeyError(
                f"Worksheet '{title}' is not registered."
            ) from exc

    def rename_sheet(
        self,
        current_title: str,
        new_title: str,
    ) -> WorksheetType:
        """Rename a registered worksheet.

        Args:
            current_title:
                Existing worksheet title.

            new_title:
                New worksheet title.

        Returns:
            The renamed worksheet.

        Raises:
            KeyError:
                If the worksheet does not exist.

            ValueError:
                If the new title is invalid or already exists.
        """
        worksheet = self.get_sheet(current_title)

        new_title = new_title.strip()

        if not new_title:
            raise ValueError("Worksheet title cannot be empty.")

        if (
            new_title != current_title
            and self.sheet_exists(new_title)
        ):
            raise ValueError(
                f"Worksheet '{new_title}' already exists."
            )

        worksheet.title = new_title

        del self._worksheet_registry[current_title]
        self._worksheet_registry[new_title] = worksheet

        self._logger.info(
            "%s Renamed worksheet '%s' -> '%s'.",
            _LOG_MESSAGE_PREFIX,
            current_title,
            new_title,
        )

        return worksheet

    def delete_sheet(
        self,
        title: str,
    ) -> None:
        """Delete a registered worksheet.

        Args:
            title:
                Worksheet title.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            KeyError:
                If the worksheet is unknown.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        worksheet = self.get_sheet(title)

        self._workbook.remove(worksheet)
        del self._worksheet_registry[title]

        self._logger.info(
            "%s Deleted worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

    def move_sheet(
        self,
        title: str,
        offset: int,
    ) -> WorksheetType:
        """Move a worksheet within the workbook.

        Args:
            title:
                Worksheet title.

            offset:
                Relative movement. Positive values move the worksheet
                toward the end of the workbook while negative values move
                it toward the beginning.

        Returns:
            The moved worksheet.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        worksheet = self.get_sheet(title)

        self._workbook.move_sheet(
            worksheet,
            offset=offset,
        )

        self._logger.debug(
            "%s Moved worksheet '%s' (offset=%d).",
            _LOG_MESSAGE_PREFIX,
            title,
            offset,
        )

        return worksheet

    def sheet_exists(
        self,
        title: str,
    ) -> bool:
        """Determine whether a worksheet is registered.

        Args:
            title:
                Worksheet title.

        Returns:
            ``True`` if the worksheet exists.
        """
        return title in self._worksheet_registry

    def list_sheets(self) -> tuple[str, ...]:
        """Return registered worksheet names.

        Returns:
            Immutable tuple preserving registry insertion order.
        """
        return tuple(self._worksheet_registry.keys())

    def hide_sheet(
        self,
        title: str,
    ) -> WorksheetType:
        """Hide a worksheet.

        Args:
            title:
                Worksheet title.

        Returns:
            The modified worksheet.
        """
        worksheet = self.get_sheet(title)

        worksheet.sheet_state = "hidden"

        self._logger.debug(
            "%s Hid worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

        return worksheet

    def unhide_sheet(
        self,
        title: str,
    ) -> WorksheetType:
        """Make a worksheet visible.

        Args:
            title:
                Worksheet title.

        Returns:
            The modified worksheet.
        """
        worksheet = self.get_sheet(title)

        worksheet.sheet_state = "visible"

        self._logger.debug(
            "%s Unhid worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

        return worksheet

    def very_hide_sheet(
        self,
        title: str,
    ) -> WorksheetType:
        """Mark a worksheet as very hidden.

        A very hidden worksheet cannot normally be made visible through the
        Excel user interface.

        Args:
            title:
                Worksheet title.

        Returns:
            The modified worksheet.
        """
        worksheet = self.get_sheet(title)

        worksheet.sheet_state = "veryHidden"

        self._logger.debug(
            "%s Very-hid worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

        return worksheet


###############################################################################
# WorkbookBuilder - Named range management
###############################################################################

from openpyxl.workbook.defined_name import DefinedName


    def create_named_range(
        self,
        name: str,
        reference: str,
        *,
        comment: str | None = None,
    ) -> DefinedName:
        """Create and register a workbook-level named range.

        If a named range having the supplied name already exists, a
        :class:`ValueError` is raised.

        Args:
            name:
                Workbook-scoped defined name.

            reference:
                Excel reference (for example ``'Teams'!$A$2:$A$100``).

            comment:
                Optional descriptive comment.

        Returns:
            The created :class:`~openpyxl.workbook.defined_name.DefinedName`.

        Raises:
            RuntimeError:
                If the workbook has not yet been created.

            ValueError:
                If the supplied arguments are invalid or the name already
                exists.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        name = name.strip()
        reference = reference.strip()

        if not name:
            raise ValueError("Named range name cannot be empty.")

        if not reference:
            raise ValueError("Named range reference cannot be empty.")

        if self.named_range_exists(name):
            raise ValueError(f"Named range '{name}' already exists.")

        defined_name = DefinedName(
            name=name,
            attr_text=reference,
            comment=comment,
        )

        self._workbook.defined_names.add(defined_name)

        self._logger.info(
            "%s Registered named range '%s'.",
            _LOG_MESSAGE_PREFIX,
            name,
        )

        return defined_name

    def update_named_range(
        self,
        name: str,
        reference: str,
    ) -> DefinedName:
        """Update the reference of an existing named range.

        Args:
            name:
                Existing defined name.

            reference:
                New Excel reference.

        Returns:
            The updated defined name.

        Raises:
            KeyError:
                If the named range does not exist.
        """
        defined_name = self.get_named_range(name)
        defined_name.attr_text = reference

        self._logger.info(
            "%s Updated named range '%s'.",
            _LOG_MESSAGE_PREFIX,
            name,
        )

        return defined_name

    def delete_named_range(
        self,
        name: str,
    ) -> None:
        """Delete a workbook-level named range.

        Args:
            name:
                Name of the defined name.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            KeyError:
                If the named range does not exist.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        if not self.named_range_exists(name):
            raise KeyError(f"Named range '{name}' does not exist.")

        del self._workbook.defined_names[name]

        self._logger.info(
            "%s Deleted named range '%s'.",
            _LOG_MESSAGE_PREFIX,
            name,
        )

    def get_named_range(
        self,
        name: str,
    ) -> DefinedName:
        """Return a workbook-level named range.

        Args:
            name:
                Defined name.

        Returns:
            Matching :class:`DefinedName`.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            KeyError:
                If the named range is not found.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        defined_name = self._workbook.defined_names.get(name)

        if defined_name is None:
            raise KeyError(f"Named range '{name}' does not exist.")

        return defined_name

    def named_range_exists(
        self,
        name: str,
    ) -> bool:
        """Determine whether a workbook named range exists.

        Args:
            name:
                Defined name.

        Returns:
            ``True`` if the defined name exists.
        """
        if self._workbook is None:
            return False

        return self._workbook.defined_names.get(name) is not None

    def register_standard_ranges(self) -> None:
        """Register application-standard workbook named ranges.

        This method is intentionally generic and does not create
        worksheet-specific ranges. Later workbook construction phases may
        extend or override this method to register ranges required by the
        cricket tournament workbook.

        If a compatible registration helper exists inside ``constants.py``,
        it will be used automatically.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Registering standard named ranges.",
            _LOG_MESSAGE_PREFIX,
        )

        register = getattr(constants, "STANDARD_NAMED_RANGES", None)

        if isinstance(register, dict):
            for name, reference in register.items():
                if not self.named_range_exists(name):
                    self.create_named_range(
                        name=name,
                        reference=str(reference),
                    )

        self._logger.info(
            "%s Standard named range registration complete.",
            _LOG_MESSAGE_PREFIX,
        )



###############################################################################
# WorkbookBuilder - Freeze pane utilities
###############################################################################

    def freeze_top_row(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Freeze the first row of a worksheet.

        Excel freezes all rows above the specified cell. Setting the freeze
        pane to ``A2`` keeps the first row visible while scrolling.

        Args:
            worksheet:
                Target worksheet.
        """
        worksheet.freeze_panes = "A2"

        self._logger.debug(
            "%s Applied top-row freeze to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def freeze_first_column(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Freeze the first column of a worksheet.

        Excel freezes all columns to the left of the specified cell. Setting
        the freeze pane to ``B1`` keeps the first column visible while
        horizontally scrolling.

        Args:
            worksheet:
                Target worksheet.
        """
        worksheet.freeze_panes = "B1"

        self._logger.debug(
            "%s Applied first-column freeze to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def freeze_custom(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
    ) -> None:
        """Apply a custom freeze pane location.

        Args:
            worksheet:
                Target worksheet.

            cell_reference:
                Excel cell reference that defines the first scrollable cell,
                such as ``B2`` or ``D5``.

        Raises:
            ValueError:
                If the supplied cell reference is empty.
        """
        reference = cell_reference.strip().upper()

        if not reference:
            raise ValueError("Cell reference cannot be empty.")

        worksheet.freeze_panes = reference

        self._logger.debug(
            "%s Applied custom freeze '%s' to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            reference,
            worksheet.title,
        )

    def clear_freeze(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Remove any freeze pane configuration.

        Args:
            worksheet:
                Target worksheet.
        """
        worksheet.freeze_panes = None

        self._logger.debug(
            "%s Cleared freeze panes for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def apply_default_freeze(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Apply the project's default freeze pane configuration.

        The default implementation freezes the first row, which is suitable for
        the majority of tabular worksheets. Subclasses may override this method
        to provide worksheet-specific behaviour while retaining a consistent
        public interface.

        Args:
            worksheet:
                Target worksheet.
        """
        self.freeze_top_row(worksheet)

        self._logger.debug(
            "%s Applied default freeze configuration to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )


###############################################################################
# WorkbookBuilder - Print configuration utilities
###############################################################################

    def set_print_area(
        self,
        worksheet: WorksheetType,
        print_area: str,
    ) -> None:
        """Configure the worksheet print area.

        Args:
            worksheet:
                Target worksheet.

            print_area:
                Excel range defining the printable area (for example
                ``A1:H50``).

        Raises:
            ValueError:
                If the supplied print area is empty.
        """
        area = print_area.strip()

        if not area:
            raise ValueError("print_area cannot be empty.")

        worksheet.print_area = area

        self._logger.debug(
            "%s Set print area for worksheet '%s' to '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            area,
        )

    def set_repeat_rows(
        self,
        worksheet: WorksheetType,
        rows: str,
    ) -> None:
        """Configure rows to repeat on every printed page.

        Args:
            worksheet:
                Target worksheet.

            rows:
                Excel row reference such as ``1:1`` or ``1:3``.

        Raises:
            ValueError:
                If the supplied row specification is empty.
        """
        repeat_rows = rows.strip()

        if not repeat_rows:
            raise ValueError("rows cannot be empty.")

        worksheet.print_title_rows = repeat_rows

        self._logger.debug(
            "%s Set repeating rows for worksheet '%s' to '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            repeat_rows,
        )

    def set_repeat_columns(
        self,
        worksheet: WorksheetType,
        columns: str,
    ) -> None:
        """Configure columns to repeat on every printed page.

        Args:
            worksheet:
                Target worksheet.

            columns:
                Excel column reference such as ``A:A`` or ``A:C``.

        Raises:
            ValueError:
                If the supplied column specification is empty.
        """
        repeat_columns = columns.strip().upper()

        if not repeat_columns:
            raise ValueError("columns cannot be empty.")

        worksheet.print_title_cols = repeat_columns

        self._logger.debug(
            "%s Set repeating columns for worksheet '%s' to '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            repeat_columns,
        )

    def set_print_titles(
        self,
        worksheet: WorksheetType,
        *,
        rows: str | None = None,
        columns: str | None = None,
    ) -> None:
        """Configure repeating print titles.

        This convenience method combines
        :meth:`set_repeat_rows` and
        :meth:`set_repeat_columns`.

        Args:
            worksheet:
                Target worksheet.

            rows:
                Optional repeating row specification.

            columns:
                Optional repeating column specification.
        """
        if rows is not None:
            self.set_repeat_rows(worksheet, rows)

        if columns is not None:
            self.set_repeat_columns(worksheet, columns)

    def set_fit_to_page(
        self,
        worksheet: WorksheetType,
        *,
        width: int | None = None,
        height: int | None = None,
    ) -> None:
        """Configure fit-to-page printing.

        When an argument is omitted, the corresponding default from
        :class:`WorkbookConfig` is used.

        Args:
            worksheet:
                Target worksheet.

            width:
                Number of pages wide.

            height:
                Number of pages tall.

        Raises:
            ValueError:
                If width or height is negative.
        """
        fit_width = (
            self._config.fit_to_width if width is None else width
        )
        fit_height = (
            self._config.fit_to_height if height is None else height
        )

        if fit_width < 0:
            raise ValueError("width cannot be negative.")

        if fit_height < 0:
            raise ValueError("height cannot be negative.")

        worksheet.page_setup.fitToPage = self._config.fit_to_page
        worksheet.page_setup.fitToWidth = fit_width
        worksheet.page_setup.fitToHeight = fit_height

        self._logger.debug(
            "%s Configured fit-to-page for worksheet '%s' "
            "(width=%d, height=%d).",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            fit_width,
            fit_height,
        )

    def set_scale(
        self,
        worksheet: WorksheetType,
        scale: int,
    ) -> None:
        """Configure print scaling.

        Args:
            worksheet:
                Target worksheet.

            scale:
                Print scaling percentage.

        Raises:
            ValueError:
                If the scale is outside Excel's supported range
                (10-400 percent).
        """
        if not 10 <= scale <= 400:
            raise ValueError(
                "scale must be between 10 and 400."
            )

        worksheet.page_setup.scale = scale

        self._logger.debug(
            "%s Set print scale for worksheet '%s' to %d%%.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            scale,
        )

    def center_on_page(
        self,
        worksheet: WorksheetType,
        *,
        horizontal: bool | None = None,
        vertical: bool | None = None,
    ) -> None:
        """Configure page centering.

        When arguments are omitted, defaults from
        :class:`WorkbookConfig` are used.

        Args:
            worksheet:
                Target worksheet.

            horizontal:
                Enable horizontal centering.

            vertical:
                Enable vertical centering.
        """
        worksheet.print_options.horizontalCentered = (
            self._config.center_horizontally
            if horizontal is None
            else horizontal
        )

        worksheet.print_options.verticalCentered = (
            self._config.center_vertically
            if vertical is None
            else vertical
        )

        self._logger.debug(
            "%s Updated page centering for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )



###############################################################################
# WorkbookBuilder - Page setup helpers
###############################################################################



    def set_page_orientation(
        self,
        worksheet: WorksheetType,
        orientation: str,
    ) -> None:
        """Set the page orientation for a worksheet.

        Args:
            worksheet:
                Target worksheet.

            orientation:
                Either ``"portrait"`` or ``"landscape"``.

        Raises:
            ValueError:
                If an unsupported orientation is supplied.
        """
        orientation = orientation.strip().lower()

        valid_orientations = {"portrait", "landscape"}

        if orientation not in valid_orientations:
            raise ValueError(
                f"Unsupported page orientation: {orientation!r}."
            )

        worksheet.page_setup.orientation = orientation

        self._logger.debug(
            "%s Set page orientation for worksheet '%s' to '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            orientation,
        )

    def set_paper_size(
        self,
        worksheet: WorksheetType,
        paper_size: int,
    ) -> None:
        """Configure the worksheet paper size.

        Args:
            worksheet:
                Target worksheet.

            paper_size:
                OpenXML paper size constant supported by ``openpyxl``.
                For example::

                    1   Letter
                    8   A3
                    9   A4
                    11  A5

        Raises:
            ValueError:
                If the paper size identifier is invalid.
        """
        if paper_size <= 0:
            raise ValueError("paper_size must be greater than zero.")

        worksheet.page_setup.paperSize = paper_size

        self._logger.debug(
            "%s Set paper size for worksheet '%s' to %d.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            paper_size,
        )

    def set_page_margins(
        self,
        worksheet: WorksheetType,
        *,
        left: float = 0.7,
        right: float = 0.7,
        top: float = 0.75,
        bottom: float = 0.75,
        header: float = 0.3,
        footer: float = 0.3,
    ) -> None:
        """Configure worksheet page margins.

        Margin values are specified in inches.

        Args:
            worksheet:
                Target worksheet.

            left:
                Left margin.

            right:
                Right margin.

            top:
                Top margin.

            bottom:
                Bottom margin.

            header:
                Header margin.

            footer:
                Footer margin.

        Raises:
            ValueError:
                If any margin is negative.
        """
        margins = {
            "left": left,
            "right": right,
            "top": top,
            "bottom": bottom,
            "header": header,
            "footer": footer,
        }

        for name, value in margins.items():
            if value < 0:
                raise ValueError(f"{name} margin cannot be negative.")

        worksheet.page_margins = PageMargins(
            left=left,
            right=right,
            top=top,
            bottom=bottom,
            header=header,
            footer=footer,
        )

        self._logger.debug(
            "%s Updated page margins for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def set_headers(
        self,
        worksheet: WorksheetType,
        *,
        left: str = "",
        center: str = "",
        right: str = "",
    ) -> None:
        """Configure worksheet page headers.

        Args:
            worksheet:
                Target worksheet.

            left:
                Left header text.

            center:
                Centre header text.

            right:
                Right header text.
        """
        worksheet.HeaderFooter = HeaderFooter()

        worksheet.oddHeader.left.text = left
        worksheet.oddHeader.center.text = center
        worksheet.oddHeader.right.text = right

        self._logger.debug(
            "%s Updated headers for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def set_footers(
        self,
        worksheet: WorksheetType,
        *,
        left: str = "",
        center: str = "",
        right: str = "",
    ) -> None:
        """Configure worksheet page footers.

        Args:
            worksheet:
                Target worksheet.

            left:
                Left footer text.

            center:
                Centre footer text.

            right:
                Right footer text.
        """
        worksheet.HeaderFooter = HeaderFooter()

        worksheet.oddFooter.left.text = left
        worksheet.oddFooter.center.text = center
        worksheet.oddFooter.right.text = right

        self._logger.debug(
            "%s Updated footers for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def set_page_numbering(
        self,
        worksheet: WorksheetType,
        *,
        first_page_number: int = 1,
        use_first_page_number: bool = True,
    ) -> None:
        """Configure worksheet page numbering.

        Args:
            worksheet:
                Target worksheet.

            first_page_number:
                Starting page number.

            use_first_page_number:
                Whether Excel should honour the supplied first page number.

        Raises:
            ValueError:
                If the first page number is less than one.
        """
        if first_page_number < 1:
            raise ValueError(
                "first_page_number must be greater than or equal to 1."
            )

        worksheet.page_setup.firstPageNumber = first_page_number
        worksheet.page_setup.useFirstPageNumber = use_first_page_number

        self._logger.debug(
            "%s Configured page numbering for worksheet '%s' "
            "(first page=%d).",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
            first_page_number,
        )



###############################################################################
# WorkbookBuilder - Protection utilities
###############################################################################

    def protect_sheet(
        self,
        worksheet: WorksheetType,
        *,
        password: str | None = None,
        select_locked_cells: bool | None = None,
        select_unlocked_cells: bool | None = None,
    ) -> None:
        """Enable worksheet protection.

        The protection options default to the values defined in
        :class:`WorkbookConfig` when not explicitly supplied.

        Args:
            worksheet:
                Target worksheet.

            password:
                Optional worksheet protection password.

            select_locked_cells:
                Whether locked cells may be selected.

            select_unlocked_cells:
                Whether unlocked cells may be selected.
        """
        protection = worksheet.protection

        protection.sheet = True

        protection.selectLockedCells = (
            self._config.allow_select_locked_cells
            if select_locked_cells is None
            else select_locked_cells
        )

        protection.selectUnlockedCells = (
            self._config.allow_select_unlocked_cells
            if select_unlocked_cells is None
            else select_unlocked_cells
        )

        if password:
            protection.set_password(password)

        self._logger.debug(
            "%s Enabled protection for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def unprotect_sheet(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Disable worksheet protection.

        Args:
            worksheet:
                Target worksheet.
        """
        worksheet.protection.sheet = False

        self._logger.debug(
            "%s Disabled protection for worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def protect_workbook(
        self,
        *,
        password: str | None = None,
        lock_structure: bool | None = None,
        lock_windows: bool | None = None,
    ) -> None:
        """Enable workbook-level protection.

        Args:
            password:
                Optional workbook password.

            lock_structure:
                Protect workbook structure.

            lock_windows:
                Protect workbook window layout.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        security = self._workbook.security

        security.lockStructure = (
            self._config.structure_protection
            if lock_structure is None
            else lock_structure
        )

        security.lockWindows = (
            self._config.windows_protection
            if lock_windows is None
            else lock_windows
        )

        if password:
            security.workbookPassword = password

        self._logger.debug(
            "%s Workbook protection enabled.",
            _LOG_MESSAGE_PREFIX,
        )

    def unlock_range(
        self,
        worksheet: WorksheetType,
        cell_range: str,
    ) -> None:
        """Unlock every cell within a worksheet range.

        Args:
            worksheet:
                Target worksheet.

            cell_range:
                Excel cell range (for example ``B2:E20``).
        """
        for row in worksheet[cell_range]:
            for cell in row:
                cell.protection = Protection(
                    locked=False,
                    hidden=cell.protection.hidden,
                )

        self._logger.debug(
            "%s Unlocked range '%s' on worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell_range,
            worksheet.title,
        )

    def lock_range(
        self,
        worksheet: WorksheetType,
        cell_range: str,
    ) -> None:
        """Lock every cell within a worksheet range.

        Args:
            worksheet:
                Target worksheet.

            cell_range:
                Excel cell range (for example ``A1:Z100``).
        """
        for row in worksheet[cell_range]:
            for cell in row:
                cell.protection = Protection(
                    locked=True,
                    hidden=cell.protection.hidden,
                )

        self._logger.debug(
            "%s Locked range '%s' on worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell_range,
            worksheet.title,
        )

    def apply_default_protection(self) -> None:
        """Apply the default workbook protection configuration.

        This helper applies workbook-level protection using the defaults
        defined in :class:`WorkbookConfig`. If worksheet protection is enabled
        by configuration, protection is also applied to all registered
        worksheets.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        if self._config.workbook_protection_enabled:
            self.protect_workbook()

        if self._config.default_sheet_protection:
            for worksheet in self._worksheet_registry.values():
                self.protect_sheet(worksheet)

        self._logger.info(
            "%s Applied default workbook protection settings.",
            _LOG_MESSAGE_PREFIX,
        )


###############################################################################
# WorkbookBuilder - Workbook metadata helpers
###############################################################################




    def set_title(
        self,
        title: str,
    ) -> None:
        """Set the workbook title.

        Args:
            title:
                Human-readable workbook title.

        Raises:
            RuntimeError:
                If the workbook has not been created.

            ValueError:
                If *title* is empty.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        title = title.strip()

        if not title:
            raise ValueError("Workbook title cannot be empty.")

        self._workbook.properties.title = title

        self._logger.debug(
            "%s Workbook title set to '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

    def set_subject(
        self,
        subject: str,
    ) -> None:
        """Set the workbook subject.

        Args:
            subject:
                Workbook subject.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.subject = subject.strip()

        self._logger.debug(
            "%s Workbook subject updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_author(
        self,
        author: str,
    ) -> None:
        """Set the workbook author.

        The value is written to both the creator and last-modified-by
        document properties.

        Args:
            author:
                Author name.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        author = author.strip()

        self._workbook.properties.creator = author
        self._workbook.properties.lastModifiedBy = author

        self._logger.debug(
            "%s Workbook author updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_manager(
        self,
        manager: str,
    ) -> None:
        """Set the workbook manager.

        Args:
            manager:
                Manager name.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.manager = manager.strip()

        self._logger.debug(
            "%s Workbook manager updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_company(
        self,
        company: str,
    ) -> None:
        """Set the workbook company.

        Args:
            company:
                Company or organization name.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.company = company.strip()

        self._logger.debug(
            "%s Workbook company updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_category(
        self,
        category: str,
    ) -> None:
        """Set the workbook category.

        Args:
            category:
                Document category.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.category = category.strip()

        self._logger.debug(
            "%s Workbook category updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_keywords(
        self,
        keywords: str,
    ) -> None:
        """Set workbook keywords.

        Args:
            keywords:
                Comma-separated keyword list.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.keywords = keywords.strip()

        self._logger.debug(
            "%s Workbook keywords updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_comments(
        self,
        comments: str,
    ) -> None:
        """Set workbook comments/description.

        Args:
            comments:
                Document description.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._workbook.properties.description = comments.strip()

        self._logger.debug(
            "%s Workbook comments updated.",
            _LOG_MESSAGE_PREFIX,
        )

    def set_creation_date(
        self,
        created: datetime | None = None,
    ) -> None:
        """Set the workbook creation timestamp.

        Args:
            created:
                Creation timestamp. When omitted, the current UTC time is used.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        timestamp = created or datetime.now(UTC)

        self._workbook.properties.created = timestamp

        self._logger.debug(
            "%s Workbook creation date set to %s.",
            _LOG_MESSAGE_PREFIX,
            timestamp.isoformat(),
        )

    def set_modification_date(
        self,
        modified: datetime | None = None,
    ) -> None:
        """Set the workbook modification timestamp.

        Args:
            modified:
                Modification timestamp. When omitted, the current UTC time is
                used.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        timestamp = modified or datetime.now(UTC)

        self._workbook.properties.modified = timestamp

        self._logger.debug(
            "%s Workbook modification date set to %s.",
            _LOG_MESSAGE_PREFIX,
            timestamp.isoformat(),
        )


###############################################################################
# WorkbookBuilder - Validation integration
###############################################################################

    def apply_standard_validations(self) -> None:
        """Apply the project's standard workbook validations.

        This method serves as the central integration point between
        ``WorkbookBuilder`` and ``validation.py``. Validation creation is
        delegated entirely to the validation module to avoid duplicating
        validation logic inside this orchestration layer.

        The implementation attempts to call one of the following public entry
        points if present in ``validation.py``:

        * ``apply_standard_validations(workbook)``
        * ``apply_standard_validations(builder)``
        * ``register_standard_validations(workbook)``
        * ``register_standard_validations(builder)``

        Missing entry points are treated as a no-op to preserve compatibility
        with evolving versions of ``validation.py``.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Applying standard workbook validations.",
            _LOG_MESSAGE_PREFIX,
        )

        self._invoke_validation_hook(
            "apply_standard_validations",
            self._workbook,
            self,
        ) or self._invoke_validation_hook(
            "register_standard_validations",
            self._workbook,
            self,
        )

        self._logger.info(
            "%s Standard validation phase completed.",
            _LOG_MESSAGE_PREFIX,
        )

    def apply_sheet_validations(
        self,
        worksheet: WorksheetType,
    ) -> None:
        """Apply validations to a single worksheet.

        Validation creation remains the responsibility of ``validation.py``.
        This method simply delegates to the first compatible interface exposed
        by that module.

        Supported public interfaces include:

        * ``apply_sheet_validations(worksheet)``
        * ``apply_sheet_validations(builder, worksheet)``
        * ``apply_sheet_validations(workbook, worksheet)``

        Args:
            worksheet:
                Worksheet receiving validations.
        """
        self._logger.debug(
            "%s Applying validations to worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

        self._invoke_validation_hook(
            "apply_sheet_validations",
            worksheet,
            self,
            self._workbook,
        )

    def register_validation_ranges(self) -> None:
        """Register workbook validation ranges.

        Workbook-level validation ranges (typically implemented as named ranges
        or hidden lookup regions) are delegated to ``validation.py``.

        Supported public interfaces include:

        * ``register_validation_ranges(workbook)``
        * ``register_validation_ranges(builder)``
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Registering validation ranges.",
            _LOG_MESSAGE_PREFIX,
        )

        self._invoke_validation_hook(
            "register_validation_ranges",
            self._workbook,
            self,
        )

    def _invoke_validation_hook(
        self,
        function_name: str,
        *preferred_arguments: object,
    ) -> bool:
        """Invoke a reusable validation hook.

        The helper searches ``validation.py`` for the requested function and
        attempts several common calling conventions. This allows
        ``WorkbookBuilder`` to remain compatible with future revisions of the
        validation subsystem without embedding validation-specific knowledge.

        Args:
            function_name:
                Name of the public function to invoke.

            *preferred_arguments:
                Candidate arguments supplied to the hook.

        Returns:
            ``True`` if a compatible validation function was successfully
            invoked; otherwise ``False``.
        """
        hook = getattr(validation, function_name, None)

        if not callable(hook):
            self._logger.debug(
                "%s Validation hook '%s' is unavailable.",
                _LOG_MESSAGE_PREFIX,
                function_name,
            )
            return False

        candidate_calls: tuple[tuple[object, ...], ...] = (
            preferred_arguments,
            (self,),
            (self._workbook,),
            (),
        )

        for args in candidate_calls:
            try:
                hook(*args)
                self._logger.debug(
                    "%s Executed validation hook '%s'.",
                    _LOG_MESSAGE_PREFIX,
                    function_name,
                )
                return True
            except TypeError:
                continue

        self._logger.warning(
            "%s Validation hook '%s' exists but no compatible signature "
            "was found.",
            _LOG_MESSAGE_PREFIX,
            function_name,
        )

        return False

    def _apply_validation_to_range(
        self,
        worksheet: WorksheetType,
        validation_object: object,
        cell_range: str,
    ) -> None:
        """Attach an existing validation object to a worksheet range.

        This helper performs no validation construction. It simply associates an
        already-created validation object with a worksheet and cell range.

        Args:
            worksheet:
                Target worksheet.

            validation_object:
                Existing validation object created by ``validation.py``.

            cell_range:
                Excel range receiving the validation.

        Raises:
            AttributeError:
                If the supplied validation object does not provide the expected
                ``add()`` interface.
        """
        validation_object.add(cell_range)
        worksheet.add_data_validation(validation_object)

        self._logger.debug(
            "%s Applied validation to '%s' on worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell_range,
            worksheet.title,
        )

    def _register_validation_object(
        self,
        key: str,
        validation_object: object,
    ) -> None:
        """Register a reusable validation object.

        Validation objects are stored inside the helper registry so that
        worksheet builders can reuse them instead of recreating identical
        validations multiple times.

        Args:
            key:
                Logical validation identifier.

            validation_object:
                Validation instance produced by ``validation.py``.
        """
        registry = self._helper_registry.setdefault(
            "validation_objects",
            {},
        )

        if isinstance(registry, dict):
            registry[key] = validation_object

        self._logger.debug(
            "%s Registered validation helper '%s'.",
            _LOG_MESSAGE_PREFIX,
            key,
        )



###############################################################################
# WorkbookBuilder - Formula integration
###############################################################################

    def register_formulas(self) -> None:
        """Register reusable formula providers.

        This method initializes the formula subsystem used throughout the
        workbook. Formula generation remains the responsibility of
        ``formulas.py``; this builder merely discovers and registers available
        helpers.

        Supported integration points include (if present):

        * ``register_formulas(builder)``
        * ``register_formulas(workbook)``
        * ``get_formula_registry()``

        Registered formula providers are cached in the builder's helper
        registry for efficient reuse.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.debug(
            "%s Registering formula providers.",
            _LOG_MESSAGE_PREFIX,
        )

        self._formula_cache.clear()

        hook = getattr(formulas, "register_formulas", None)

        if callable(hook):
            try:
                hook(self)
            except TypeError:
                hook(self._workbook)

        registry_factory = getattr(
            formulas,
            "get_formula_registry",
            None,
        )

        if callable(registry_factory):
            registry = registry_factory()

            if isinstance(registry, dict):
                self._formula_cache.update(registry)

        self._logger.info(
            "%s Formula registration completed.",
            _LOG_MESSAGE_PREFIX,
        )

    def insert_formula(
        self,
        cell: CellType,
        formula: str,
    ) -> None:
        """Insert an Excel formula into a cell.

        Args:
            cell:
                Destination cell.

            formula:
                Excel formula.

        Raises:
            ValueError:
                If *formula* is empty.
        """
        formula = formula.strip()

        if not formula:
            raise ValueError("Formula cannot be empty.")

        if not formula.startswith("="):
            formula = f"={formula}"

        cell.value = formula

        self._logger.debug(
            "%s Inserted formula into %s.",
            _LOG_MESSAGE_PREFIX,
            cell.coordinate,
        )

    def insert_registered_formula(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
        formula_name: str,
        *args: object,
        **kwargs: object,
    ) -> None:
        """Insert a dynamically generated formula.

        The formula is obtained from the registered formula cache or, if
        necessary, lazily resolved from ``formulas.py``.

        Args:
            worksheet:
                Target worksheet.

            cell_reference:
                Destination cell reference.

            formula_name:
                Logical formula identifier.

            *args:
                Positional arguments forwarded to the formula provider.

            **kwargs:
                Keyword arguments forwarded to the formula provider.

        Raises:
            KeyError:
                If the requested formula provider cannot be located.
        """
        provider = self.get_formula_provider(formula_name)

        if callable(provider):
            expression = provider(*args, **kwargs)
        else:
            expression = provider

        self.insert_formula(
            worksheet[cell_reference],
            str(expression),
        )

    def get_formula_provider(
        self,
        formula_name: str,
    ) -> object:
        """Return a registered formula provider.

        Formula providers are cached after first discovery to avoid repeated
        module lookups.

        Args:
            formula_name:
                Logical provider name.

        Returns:
            The registered provider.

        Raises:
            KeyError:
                If the provider cannot be found.
        """
        if formula_name in self._formula_cache:
            return self._formula_cache[formula_name]

        provider = getattr(formulas, formula_name, None)

        if provider is None:
            raise KeyError(
                f"Unknown formula provider: {formula_name}"
            )

        self._formula_cache[formula_name] = provider

        return provider

    def cache_formula(
        self,
        name: str,
        provider: object,
    ) -> None:
        """Register a reusable formula provider.

        Args:
            name:
                Logical formula name.

            provider:
                Formula callable or constant expression.
        """
        self._formula_cache[name] = provider

        self._logger.debug(
            "%s Cached formula provider '%s'.",
            _LOG_MESSAGE_PREFIX,
            name,
        )

    def clear_formula_cache(self) -> None:
        """Clear all cached formula providers."""
        self._formula_cache.clear()

        self._logger.debug(
            "%s Formula cache cleared.",
            _LOG_MESSAGE_PREFIX,
        )

    def list_registered_formulas(self) -> tuple[str, ...]:
        """Return the names of all cached formula providers.

        Returns:
            Alphabetically sorted tuple of registered provider names.
        """
        return tuple(sorted(self._formula_cache))

    def _initialize_formula_helpers(self) -> None:
        """Initialize internal formula infrastructure.

        This helper prepares the cache used for dynamic formula discovery.
        Calling this method multiple times is safe.
        """
        if not hasattr(self, "_formula_cache"):
            self._formula_cache: dict[str, object] = {}

        self._helper_registry["formula_cache"] = self._formula_cache

        self._logger.debug(
            "%s Formula helper infrastructure initialized.",
            _LOG_MESSAGE_PREFIX,
        )


###############################################################################
# WorkbookBuilder - Hyperlink helpers
###############################################################################



    def add_internal_hyperlink(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
        target: str,
        *,
        display: str | None = None,
        tooltip: str | None = None,
    ) -> CellType:
        """Create an internal workbook hyperlink.

        Internal hyperlinks navigate to another worksheet or cell within the
        same workbook.

        Args:
            worksheet:
                Worksheet containing the hyperlink.

            cell_reference:
                Cell that will contain the hyperlink.

            target:
                Excel internal reference such as::

                    Summary!A1
                    'Match 01'!B5

            display:
                Optional displayed text.

            tooltip:
                Optional Excel tooltip.

        Returns:
            The modified cell.
        """
        cell = worksheet[cell_reference]

        if display is not None:
            cell.value = display

        cell.hyperlink = Hyperlink(
            ref=cell.coordinate,
            location=target,
            tooltip=tooltip,
            display=display,
        )

        try:
            self.style_cell(cell, "Hyperlink")
        except Exception:
            cell.font = Font(
                color="0563C1",
                underline="single",
            )

        self._logger.debug(
            "%s Added internal hyperlink '%s' -> '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell.coordinate,
            target,
        )

        return cell

    def add_external_hyperlink(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
        url: str,
        *,
        display: str | None = None,
        tooltip: str | None = None,
    ) -> CellType:
        """Create an external hyperlink.

        Args:
            worksheet:
                Target worksheet.

            cell_reference:
                Destination cell.

            url:
                External URL.

            display:
                Optional displayed text.

            tooltip:
                Optional tooltip.

        Returns:
            The modified cell.
        """
        cell = worksheet[cell_reference]

        if display is not None:
            cell.value = display

        cell.hyperlink = Hyperlink(
            ref=cell.coordinate,
            target=url,
            tooltip=tooltip,
            display=display,
        )

        try:
            self.style_cell(cell, "Hyperlink")
        except Exception:
            cell.font = Font(
                color="0563C1",
                underline="single",
            )

        self._logger.debug(
            "%s Added external hyperlink '%s'.",
            _LOG_MESSAGE_PREFIX,
            url,
        )

        return cell

    def add_sheet_link(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
        target_sheet: str,
        *,
        target_cell: str = "A1",
        display: str | None = None,
    ) -> CellType:
        """Create a hyperlink to another worksheet.

        Args:
            worksheet:
                Source worksheet.

            cell_reference:
                Cell containing the hyperlink.

            target_sheet:
                Destination worksheet title.

            target_cell:
                Destination cell.

            display:
                Optional displayed text.

        Returns:
            The modified cell.

        Raises:
            KeyError:
                If the destination worksheet is not registered.
        """
        self.get_sheet(target_sheet)

        destination = f"'{target_sheet}'!{target_cell}"

        return self.add_internal_hyperlink(
            worksheet=worksheet,
            cell_reference=cell_reference,
            target=destination,
            display=display or target_sheet,
        )

    def create_table_of_contents_links(
        self,
        worksheet: WorksheetType,
        *,
        start_row: int = 2,
        column: int | str = 1,
        include_hidden: bool = False,
    ) -> None:
        """Populate a worksheet with links to registered worksheets.

        Each worksheet is written to one row beginning at *start_row*.

        Args:
            worksheet:
                Worksheet acting as the table of contents.

            start_row:
                First output row.

            column:
                Output column.

            include_hidden:
                Include hidden worksheets.
        """
        column_letter = self.validate_column_reference(column)

        row = start_row

        for sheet_name, sheet in self._worksheet_registry.items():
            if (
                not include_hidden
                and sheet.sheet_state != "visible"
            ):
                continue

            self.add_sheet_link(
                worksheet=worksheet,
                cell_reference=f"{column_letter}{row}",
                target_sheet=sheet_name,
                display=sheet_name,
            )

            row += 1

        self._logger.info(
            "%s Created table of contents links.",
            _LOG_MESSAGE_PREFIX,
        )

    def add_navigation_button(
        self,
        worksheet: WorksheetType,
        cell_reference: str,
        *,
        target_sheet: str,
        label: str,
        target_cell: str = "A1",
    ) -> CellType:
        """Create a simple worksheet navigation button.

        The implementation intentionally uses a styled hyperlink instead of
        Excel drawing objects to maximize compatibility with openpyxl and
        spreadsheet applications.

        Args:
            worksheet:
                Source worksheet.

            cell_reference:
                Cell representing the navigation button.

            target_sheet:
                Destination worksheet.

            label:
                Button caption.

            target_cell:
                Destination cell.

        Returns:
            The modified cell.
        """
        cell = self.add_sheet_link(
            worksheet=worksheet,
            cell_reference=cell_reference,
            target_sheet=target_sheet,
            target_cell=target_cell,
            display=label,
        )

        try:
            self.style_cell(cell, "Navigation")
        except Exception:
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

        self._logger.debug(
            "%s Added navigation button '%s' -> '%s'.",
            _LOG_MESSAGE_PREFIX,
            label,
            target_sheet,
        )

        return cell



###############################################################################
# WorkbookBuilder - Workbook navigation helpers
###############################################################################

    def create_main_menu_sheet(
        self,
        *,
        title: str = "Main Menu",
        overwrite: bool = False,
    ) -> WorksheetType:
        """Create the workbook's main navigation worksheet.

        The main menu serves as the central entry point into the workbook and
        typically contains hyperlinks to every user-facing worksheet.

        Args:
            title:
                Title of the navigation worksheet.

            overwrite:
                If ``True`` and the worksheet already exists, the existing
                worksheet is deleted before creating a new one.

        Returns:
            The created or existing worksheet.

        Raises:
            RuntimeError:
                If the workbook has not been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        if self.sheet_exists(title):
            if overwrite:
                self.delete_sheet(title)
            else:
                return self.get_sheet(title)

        worksheet = self.create_sheet(title, index=0)

        worksheet["A1"] = self._config.workbook_title
        worksheet["A2"] = "Workbook Navigation"

        try:
            self.style_cell(worksheet["A1"], "Title")
            self.style_cell(worksheet["A2"], "Header")
        except Exception:
            pass

        self._logger.info(
            "%s Created main menu worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            title,
        )

        return worksheet

    def build_navigation_table(
        self,
        worksheet: WorksheetType,
        *,
        start_row: int = 4,
        title: str = "Available Worksheets",
        include_hidden: bool = False,
    ) -> None:
        """Build a navigation table containing worksheet hyperlinks.

        Args:
            worksheet:
                Worksheet that will contain the navigation table.

            start_row:
                First row used by the table.

            title:
                Section title.

            include_hidden:
                Whether hidden worksheets should be listed.
        """
        worksheet.cell(row=start_row, column=1).value = title

        try:
            self.style_cell(
                worksheet.cell(row=start_row, column=1),
                "Header",
            )
        except Exception:
            pass

        current_row = start_row + 1

        for name, sheet in self._worksheet_registry.items():
            if (
                not include_hidden
                and sheet.sheet_state != "visible"
            ):
                continue

            if sheet is worksheet:
                continue

            self.add_sheet_link(
                worksheet=worksheet,
                cell_reference=f"A{current_row}",
                target_sheet=name,
                display=name,
            )

            current_row += 1

        self._logger.debug(
            "%s Built navigation table on worksheet '%s'.",
            _LOG_MESSAGE_PREFIX,
            worksheet.title,
        )

    def add_back_link(
        self,
        worksheet: WorksheetType,
        *,
        target_sheet: str,
        target_cell: str = "A1",
        cell_reference: str = "A1",
        label: str = "← Back",
    ) -> CellType:
        """Insert a reusable back-navigation hyperlink.

        Args:
            worksheet:
                Worksheet receiving the hyperlink.

            target_sheet:
                Destination worksheet.

            target_cell:
                Destination cell.

            cell_reference:
                Cell containing the hyperlink.

            label:
                Display text.

        Returns:
            The modified cell.
        """
        return self.add_navigation_button(
            worksheet=worksheet,
            cell_reference=cell_reference,
            target_sheet=target_sheet,
            target_cell=target_cell,
            label=label,
        )

    def add_home_link(
        self,
        worksheet: WorksheetType,
        *,
        home_sheet: str = "Main Menu",
        cell_reference: str = "B1",
        label: str = "🏠 Home",
    ) -> CellType:
        """Insert a hyperlink returning to the workbook home page.

        Args:
            worksheet:
                Worksheet receiving the hyperlink.

            home_sheet:
                Name of the home worksheet.

            cell_reference:
                Cell containing the hyperlink.

            label:
                Display text.

        Returns:
            The modified cell.
        """
        return self.add_navigation_button(
            worksheet=worksheet,
            cell_reference=cell_reference,
            target_sheet=home_sheet,
            label=label,
        )

    def create_sheet_index(
        self,
        *,
        sheet_name: str = "Sheet Index",
        include_hidden: bool = False,
    ) -> WorksheetType:
        """Create an alphabetical worksheet index.

        Unlike the main menu, the sheet index is intended to provide a compact
        directory of workbook worksheets.

        Args:
            sheet_name:
                Worksheet title.

            include_hidden:
                Include hidden worksheets.

        Returns:
            The worksheet containing the index.
        """
        if self.sheet_exists(sheet_name):
            worksheet = self.get_sheet(sheet_name)
        else:
            worksheet = self.create_sheet(sheet_name)

        worksheet["A1"] = "Worksheet Index"

        try:
            self.style_cell(worksheet["A1"], "Title")
        except Exception:
            pass

        row = 3

        for name in sorted(self._worksheet_registry):
            sheet = self._worksheet_registry[name]

            if (
                not include_hidden
                and sheet.sheet_state != "visible"
            ):
                continue

            if name == sheet_name:
                continue

            self.add_sheet_link(
                worksheet=worksheet,
                cell_reference=f"A{row}",
                target_sheet=name,
                display=name,
            )

            row += 1

        self._logger.info(
            "%s Created worksheet index '%s'.",
            _LOG_MESSAGE_PREFIX,
            sheet_name,
        )

        return worksheet


###############################################################################
# WorkbookBuilder - General helper utilities
###############################################################################


    def write_cell(
        self,
        worksheet: WorksheetType,
        coordinate: str,
        value: object,
        *,
        style: str | None = None,
    ) -> CellType:
        """Safely write a value to a worksheet cell.

        This helper validates that the destination is not a merged-cell
        placeholder before assigning the value. Optionally, a named style may
        be applied after writing.

        Args:
            worksheet:
                Target worksheet.

            coordinate:
                Excel cell coordinate.

            value:
                Value to write.

            style:
                Optional named style.

        Returns:
            The modified cell.

        Raises:
            ValueError:
                If the destination cell is a merged-cell placeholder.
        """
        cell = worksheet[coordinate]

        if isinstance(cell, MergedCell):
            raise ValueError(
                f"Cannot write directly to merged cell '{coordinate}'."
            )

        cell.value = value

        if style is not None:
            self.style_cell(cell, style)

        return cell

    def merge_cells(
        self,
        worksheet: WorksheetType,
        cell_range: str,
        *,
        value: object | None = None,
        style: str | None = None,
    ) -> CellType:
        """Merge a range of cells.

        Optionally assigns a value and style to the upper-left cell.

        Args:
            worksheet:
                Target worksheet.

            cell_range:
                Excel range to merge.

            value:
                Optional value for the anchor cell.

            style:
                Optional named style.

        Returns:
            The upper-left cell of the merged range.
        """
        worksheet.merge_cells(cell_range)

        first_coordinate = cell_range.split(":")[0]
        cell = worksheet[first_coordinate]

        if value is not None:
            cell.value = value

        if style is not None:
            self.style_cell(cell, style)

        self._logger.debug(
            "%s Merged range '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell_range,
        )

        return cell

    def unmerge_cells(
        self,
        worksheet: WorksheetType,
        cell_range: str,
    ) -> None:
        """Unmerge a merged cell range.

        Args:
            worksheet:
                Target worksheet.

            cell_range:
                Excel merged range.
        """
        worksheet.unmerge_cells(cell_range)

        self._logger.debug(
            "%s Unmerged range '%s'.",
            _LOG_MESSAGE_PREFIX,
            cell_range,
        )

    def parse_range(
        self,
        cell_range: str,
    ) -> tuple[int, int, int, int]:
        """Parse an Excel range into numeric boundaries.

        Args:
            cell_range:
                Excel range.

        Returns:
            Tuple containing::

                (
                    min_column,
                    min_row,
                    max_column,
                    max_row,
                )
        """
        return range_boundaries(cell_range)

    def iter_range_coordinates(
        self,
        cell_range: str,
    ) -> Iterable[str]:
        """Iterate over every coordinate within a cell range.

        Args:
            cell_range:
                Excel range.

        Yields:
            Cell coordinates in row-major order.
        """
        for row in rows_from_range(cell_range):
            yield from row

    def split_coordinate(
        self,
        coordinate: str,
    ) -> tuple[str, int]:
        """Split a coordinate into column letter and row number.

        Args:
            coordinate:
                Excel coordinate.

        Returns:
            Tuple of column letter and row number.
        """
        return coordinate_from_string(coordinate)

    def make_coordinate(
        self,
        column: int | str,
        row: int,
    ) -> str:
        """Construct an Excel coordinate.

        Args:
            column:
                Column index or letter.

            row:
                Row number.

        Returns:
            Excel coordinate.
        """
        column_letter = (
            get_column_letter(column)
            if isinstance(column, int)
            else self.validate_column_reference(column)
        )

        return f"{column_letter}{row}"

    def column_to_index(
        self,
        column: str,
    ) -> int:
        """Convert an Excel column letter to a numeric index.

        Args:
            column:
                Excel column letter.

        Returns:
            One-based column index.
        """
        return column_index_from_string(column.upper())

    def worksheet_dimensions(
        self,
        worksheet: WorksheetType,
    ) -> tuple[int, int]:
        """Return worksheet dimensions.

        Args:
            worksheet:
                Target worksheet.

        Returns:
            Tuple containing::

                (
                    maximum_row,
                    maximum_column,
                )
        """
        return (
            worksheet.max_row,
            worksheet.max_column,
        )

    def used_range(
        self,
        worksheet: WorksheetType,
    ) -> str:
        """Return the worksheet's current used range.

        Args:
            worksheet:
                Target worksheet.

        Returns:
            Excel range string.
        """
        return worksheet.calculate_dimension()

    def log_method_call(
        self,
        method_name: str,
        **context: object,
    ) -> None:
        """Emit a standardized debug log entry.

        Args:
            method_name:
                Method being logged.

            **context:
                Additional contextual information.
        """
        if context:
            details = ", ".join(
                f"{key}={value!r}"
                for key, value in sorted(context.items())
            )
            message = f"{method_name}({details})"
        else:
            message = method_name

        self._logger.debug(
            "%s %s",
            _LOG_MESSAGE_PREFIX,
            message,
        )

    @contextmanager
    def performance_timer(
        self,
        operation: str,
    ) -> Iterable[None]:
        """Measure the execution time of an operation.

        Example:
            >>> with builder.performance_timer("Populate Teams"):
            ...     ...

        Args:
            operation:
                Human-readable operation name.
        """
        start = perf_counter()

        try:
            yield
        finally:
            elapsed = perf_counter() - start

            self._logger.info(
                "%s %s completed in %.6f seconds.",
                _LOG_MESSAGE_PREFIX,
                operation,
                elapsed,
            )

    def batch_write(
        self,
        worksheet: WorksheetType,
        values: Sequence[tuple[str, object]],
        *,
        style: str | None = None,
    ) -> None:
        """Efficiently write multiple values.

        Args:
            worksheet:
                Target worksheet.

            values:
                Sequence of ``(coordinate, value)`` pairs.

            style:
                Optional style applied to every written cell.
        """
        for coordinate, value in values:
            self.write_cell(
                worksheet,
                coordinate,
                value,
                style=style,
            )

        self._logger.debug(
            "%s Batch wrote %d cells.",
            _LOG_MESSAGE_PREFIX,
            len(values),
        )


###############################################################################
# WorkbookBuilder - Public API and convenience interface
###############################################################################


from openpyxl import load_workbook as _load_workbook


    def build_workbook(self) -> Workbook:
        """Build and initialize a workbook.

        This is the primary public entry point for workbook construction.
        The method orchestrates all initialization phases implemented by
        previous builder methods.

        The workflow intentionally delegates to reusable helper methods so
        subclasses may override individual phases without changing the public
        API.

        Returns:
            The fully initialized workbook instance.

        Raises:
            RuntimeError:
                If workbook creation fails.
        """
        self._logger.info(
            "%s Starting workbook build.",
            _LOG_MESSAGE_PREFIX,
        )

        if self._workbook is None:
            self.create_workbook()

        self.initialize_workbook()
        self.initialize_internal_state()

        try:
            self.initialize_styles()
        except Exception:
            self._logger.exception(
                "%s Style initialization failed.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

        try:
            self.register_standard_ranges()
        except Exception:
            self._logger.exception(
                "%s Named range registration failed.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

        try:
            self.apply_standard_validations()
        except Exception:
            self._logger.exception(
                "%s Validation initialization failed.",
                _LOG_MESSAGE_PREFIX,
            )
            raise

        self._logger.info(
            "%s Workbook build completed.",
            _LOG_MESSAGE_PREFIX,
        )

        return self.workbook

    def save_workbook(
        self,
        filename: str | Path | BinaryIO,
    ) -> None:
        """Save the workbook.

        Args:
            filename:
                File path or writable binary file object.

        Raises:
            RuntimeError:
                If no workbook has been created.
        """
        if self._workbook is None:
            raise RuntimeError("Workbook has not been created.")

        self._logger.info(
            "%s Saving workbook to %s.",
            _LOG_MESSAGE_PREFIX,
            filename,
        )

        self._workbook.save(filename)

    def load_workbook(
        self,
        filename: str | Path,
        *,
        read_only: bool = False,
        keep_vba: bool = False,
        data_only: bool = False,
    ) -> Workbook:
        """Load an existing workbook.

        The worksheet registry is automatically rebuilt after loading.

        Args:
            filename:
                Workbook filename.

            read_only:
                Open workbook in read-only mode.

            keep_vba:
                Preserve VBA content.

            data_only:
                Read cached formula values.

        Returns:
            Loaded workbook.
        """
        self._logger.info(
            "%s Loading workbook %s.",
            _LOG_MESSAGE_PREFIX,
            filename,
        )

        self._workbook = _load_workbook(
            filename=filename,
            read_only=read_only,
            keep_vba=keep_vba,
            data_only=data_only,
        )

        self._worksheet_registry.clear()

        for worksheet in self._workbook.worksheets:
            self._worksheet_registry[worksheet.title] = worksheet

        return self._workbook

    def export_workbook(
        self,
        filename: str | Path,
    ) -> Path:
        """Build and export a workbook.

        This convenience method combines workbook construction and saving.

        Args:
            filename:
                Destination filename.

        Returns:
            Normalized output path.
        """
        output = Path(filename)

        self.build_workbook()
        self.save_workbook(output)

        self._logger.info(
            "%s Workbook exported to %s.",
            _LOG_MESSAGE_PREFIX,
            output,
        )

        return output

    def __enter__(self) -> Self:
        """Enter the builder context.

        Returns:
            The builder instance.
        """
        if self._workbook is None:
            self.create_workbook()

        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: object | None,
    ) -> bool:
        """Exit the builder context.

        Args:
            exc_type:
                Exception type.

            exc_value:
                Exception instance.

            traceback:
                Traceback object.

        Returns:
            ``False`` so any exception is propagated.
        """
        if exc_type is not None:
            self._logger.exception(
                "%s Workbook builder exited with an exception.",
                _LOG_MESSAGE_PREFIX,
                exc_info=(exc_type, exc_value, traceback),
            )

        return False


###############################################################################
# Convenience functions
###############################################################################

def build_workbook(
    config: WorkbookConfig | None = None,
) -> Workbook:
    """Build a workbook using the default builder.

    Args:
        config:
            Optional workbook configuration.

    Returns:
        Newly built workbook.
    """
    builder = WorkbookBuilder(config or WorkbookConfig())
    return builder.build_workbook()


def create_workbook_builder(
    config: WorkbookConfig | None = None,
) -> WorkbookBuilder:
    """Create a configured workbook builder.

    Args:
        config:
            Optional workbook configuration.

    Returns:
        Configured :class:`WorkbookBuilder`.
    """
    return WorkbookBuilder(config or WorkbookConfig())


###############################################################################
# Public exports
###############################################################################

__all__ = [
    "WorkbookConfig",
    "WorkbookBuilder",
    "build_workbook",
    "create_workbook_builder",
]

###############################################################################
# Module finalization
###############################################################################
"""
Workbook Module Finalization
============================

This section completes the public surface of ``workbook.py``.

Design Goals
------------
The module intentionally acts as the orchestration layer of the project.

Responsibilities include:

* workbook lifecycle management
* worksheet lifecycle management
* style integration
* validation integration
* formula integration
* workbook metadata
* workbook navigation
* workbook protection
* reusable workbook utilities

Responsibilities intentionally excluded from this module:

* cricket-specific business rules
* scoring logic
* validation implementations
* formula implementations
* style definitions
* application constants

Those responsibilities remain delegated to their dedicated project modules.

Module Compatibility
--------------------
This module has been designed to integrate with:

* ``constants.py`` — application constants and configuration values
* ``styles.py`` — reusable NamedStyle registration and styling helpers
* ``validation.py`` — reusable validation builders
* ``formulas.py`` — reusable Excel formula helpers

No business logic from those modules should be duplicated here.

Public API
----------
The supported public API consists of:

* WorkbookConfig
* WorkbookBuilder
* build_workbook()
* create_workbook_builder()

Everything else should be considered an implementation detail unless
explicitly documented otherwise.

Python Compatibility
--------------------
Designed for:

* Python 3.12+
* openpyxl 3.1+

Implementation Notes
--------------------
WorkbookBuilder intentionally centralizes orchestration while delegating
specialized functionality to the corresponding project modules. This
keeps responsibilities well separated and minimizes maintenance overhead
as the project grows.
"""

###############################################################################
# Public API verification
###############################################################################

# Normalize exports to ensure a stable public interface.
__all__ = tuple(
    dict.fromkeys(
        (
            "WorkbookConfig",
            "WorkbookBuilder",
            "build_workbook",
            "create_workbook_builder",
        )
    )
)

###############################################################################
# Compatibility verification metadata
###############################################################################

# These module references are intentionally retained so static analysis tools
# and future maintenance can easily verify the expected integration points.
_MODULE_COMPATIBILITY: Final[tuple[str, ...]] = (
    "constants",
    "styles",
    "validation",
    "formulas",
)

###############################################################################
# End of workbook.py
###############################################################################